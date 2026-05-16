"""Excel-rapportgenerator voor de pensioenprognose."""

from __future__ import annotations

import io
from datetime import date
from decimal import Decimal
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from pensioen.models.cashflow import HuishoudCashflow, JaarResultaat
from pensioen.calculations.scenario_engine import ScenarioVergelijking

# Kleurenpalet
KLEUR_HEADER = "1F4E79"  # donkerblauw
KLEUR_SUBHEADER = "BDD7EE"  # lichtblauw
KLEUR_TEKORT = "FFC7CE"  # lichtroze (tekortjaar)
KLEUR_POSITIEF = "C6EFCE"  # lichtgroen

_EURO = "€ #,##0"
_PCT = "0.00%"


def _header_stijl(cel: openpyxl.cell.Cell, donker: bool = True) -> None:
    kleur = KLEUR_HEADER if donker else KLEUR_SUBHEADER
    tekst_kleur = "FFFFFF" if donker else "000000"
    cel.fill = PatternFill("solid", fgColor=kleur)
    cel.font = Font(bold=True, color=tekst_kleur)
    cel.alignment = Alignment(horizontal="center")


def _kolom_breedte(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(
            max(max_len + 2, 10), 30
        )


def _schrijf_jaaroverzicht(wb: openpyxl.Workbook, cashflow: HuishoudCashflow) -> None:
    """Tab 1: Jaaroverzicht met alle jaren op één regel."""
    ws = wb.create_sheet("Jaaroverzicht")
    headers = [
        "Jaar",
        "Arbeid bruto",
        "AOW bruto",
        "Pensioen bruto",
        "Totaal bruto",
        "Belasting",
        "Heffingskorting",
        "Totaal netto",
        "Netto p/m",
        "Eff. tarief %",
        "Vermogen einde jaar",
        "Tekortjaar?",
    ]
    for col, header in enumerate(headers, 1):
        cel = ws.cell(row=1, column=col, value=header)
        _header_stijl(cel, donker=True)

    for rij, jr in enumerate(cashflow.jaren, 2):
        waarden = [
            jr.jaar,
            float(jr.arbeid_bruto),
            float(jr.aow_bruto),
            float(jr.pensioen_bruto),
            float(jr.totaal_bruto),
            float(jr.totaal_belasting),
            float(jr.totaal_heffingskorting),
            float(jr.netto),
            float(jr.netto_per_maand),
            float(jr.effectief_tarief / Decimal("100")),
            float(jr.vermogen_einde_jaar),
            "JA" if jr.is_tekortjaar else "nee",
        ]
        for col, waarde in enumerate(waarden, 1):
            cel = ws.cell(row=rij, column=col, value=waarde)
            if col in (2, 3, 4, 5, 6, 7, 8, 9, 11) and isinstance(waarde, float):
                cel.number_format = _EURO
            elif col == 10:
                cel.number_format = _PCT
        if jr.is_tekortjaar:
            for col in range(1, len(headers) + 1):
                ws.cell(row=rij, column=col).fill = PatternFill(
                    "solid", fgColor=KLEUR_TEKORT
                )

    _kolom_breedte(ws)


def _schrijf_maanddetail(wb: openpyxl.Workbook, cashflow: HuishoudCashflow) -> None:
    """Tab 2: Maanddetail (alle maanden van alle jaren)."""
    ws = wb.create_sheet("Maanddetail")
    headers = [
        "Jaar", "Maand",
        "Arbeid P1", "Arbeid P2",
        "AOW P1", "AOW P2",
        "Pensioen P1", "Pensioen P2",
        "Rente", "Incidenteel ontvangst", "Incidenteel uitgave",
        "Belasting P1", "HK P1", "Belasting P2", "HK P2",
        "Vermogen einde maand", "Netto",
    ]
    for col, header in enumerate(headers, 1):
        cel = ws.cell(row=1, column=col, value=header)
        _header_stijl(cel, donker=True)

    rij = 2
    for jr in cashflow.jaren:
        for m in jr.maanden:
            waarden = [
                m.jaar, m.maand,
                float(m.arbeid_p1_bruto), float(m.arbeid_p2_bruto),
                float(m.aow_p1_bruto), float(m.aow_p2_bruto),
                float(m.pensioen_p1_bruto), float(m.pensioen_p2_bruto),
                float(m.rente_bruto),
                float(m.eenmalig_ontvangst), float(m.eenmalig_uitgave),
                float(m.belasting_p1), float(m.heffingskorting_p1),
                float(m.belasting_p2), float(m.heffingskorting_p2),
                float(m.vermogen_einde_maand),
                float(m.netto),
            ]
            for col, waarde in enumerate(waarden, 1):
                cel = ws.cell(row=rij, column=col, value=waarde)
                if col > 2 and isinstance(waarde, float):
                    cel.number_format = _EURO
            rij += 1

    _kolom_breedte(ws)


def _schrijf_aannames(wb: openpyxl.Workbook, cashflow: HuishoudCashflow) -> None:
    """Tab 3: Aannames en belastingkentallen."""
    ws = wb.create_sheet("Aannames")
    ws.cell(row=1, column=1, value="Aannames en disclaimers").font = Font(bold=True, size=12)
    ws.cell(row=2, column=1, value=f"Scenario: {cashflow.scenario_naam}")
    ws.cell(row=3, column=1, value=f"Gegenereerd op: {date.today()}")

    rij = 5
    ws.cell(row=rij, column=1, value="Melding").font = Font(bold=True)
    rij += 1
    for aanname in cashflow.aannames:
        ws.cell(row=rij, column=1, value=aanname)
        rij += 1

    rij += 1
    ws.cell(row=rij, column=1, value="Gebruikte tariefsjaren per prognosejaar").font = Font(bold=True)
    rij += 1
    for jr in cashflow.jaren:
        ws.cell(row=rij, column=1, value=jr.jaar)
        ws.cell(row=rij, column=2, value=f"Tariefsjaar {jr.tarieven_jaar}")
        if jr.tarieven_aanname:
            ws.cell(row=rij, column=3, value=f"⚠ {jr.tarieven_aanname}")
        rij += 1

    _kolom_breedte(ws)


def _schrijf_vergelijking(
    wb: openpyxl.Workbook, vergelijking: ScenarioVergelijking
) -> None:
    """Tab 4: Scenariovergelijking."""
    ws = wb.create_sheet("Vergelijking")
    headers = [
        "Scenario", "Stopdatum werk",
        "Mediaan netto p/m", "Netto laagste jaar",
        "Laagste inkomensjaar", "Vermogen op 70",
        "Vermogen op 80", "Gem. belastingdruk %",
        "Aantal tekortjaren",
    ]
    for col, header in enumerate(headers, 1):
        _header_stijl(ws.cell(row=1, column=col, value=header), donker=True)

    for rij, sr in enumerate(vergelijking.scenario_resultaten, 2):
        waarden = [
            sr.scenario_naam, sr.stopdatum_werk,
            float(sr.netto_per_maand_mediaan), float(sr.netto_laagste_jaar),
            sr.laagste_inkomensjaar or "",
            float(sr.vermogen_op_70), float(sr.vermogen_op_80),
            float(sr.gemiddelde_belastingdruk / Decimal("100")),
            sr.aantal_tekortjaren,
        ]
        for col, waarde in enumerate(waarden, 1):
            cel = ws.cell(row=rij, column=col, value=waarde)
            if col in (3, 4, 6, 7) and isinstance(waarde, float):
                cel.number_format = _EURO
            elif col == 8 and isinstance(waarde, float):
                cel.number_format = _PCT

    _kolom_breedte(ws)


def genereer_rapport(
    cashflow: HuishoudCashflow,
    vergelijking: ScenarioVergelijking | None = None,
) -> bytes:
    """
    Genereer een volledig Excel-rapport en geef het terug als bytes.

    Args:
        cashflow: Berekeningsresultaat van het hoofdscenario.
        vergelijking: Optionele scenariovergelijking.

    Returns:
        Excel-bestand als bytes (kan direct worden aangeboden als download).
    """
    wb = openpyxl.Workbook()
    # Verwijder het standaard lege werkblad
    wb.remove(wb.active)

    _schrijf_jaaroverzicht(wb, cashflow)
    _schrijf_maanddetail(wb, cashflow)
    _schrijf_aannames(wb, cashflow)
    if vergelijking:
        _schrijf_vergelijking(wb, vergelijking)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
