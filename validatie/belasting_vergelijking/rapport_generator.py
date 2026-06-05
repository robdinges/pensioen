"""Rapport generatie voor belastingvergelijkingen.

Genereert markdown en optioneel Excel rapporten van verschillenanalyses.
"""

from __future__ import annotations

from datetime import datetime
from decimal import Decimal
from io import BytesIO
from pathlib import Path

from .vergelijker import VergelijkingsResultaat, VerschilAnalyse


def genereer_markdown_rapport(resultaat: VergelijkingsResultaat) -> str:
    """
    Genereer markdown rapport van vergelijking.
    
    Args:
        resultaat: VergelijkingsResultaat van vergelijking
    
    Returns:
        Markdown string
    """
    lines = []
    
    # Header
    lines.append(f"# Belastingvergelijking: dutch_tax vs pensioen-app")
    lines.append("")
    lines.append(f"**Huishouden**: {resultaat.huishouden_id}")
    lines.append(f"**Submission jaar**: {resultaat.submission_jaar}")
    lines.append(f"**Berekening jaar**: {resultaat.jaar}")
    lines.append(f"**Gegenereerd**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append("")
    lines.append("---")
    lines.append("")
    
    # Samenvatting
    lines.append("## 📊 Samenvatting")
    lines.append("")
    lines.append(f"- **Totaal verschillen**: {len(resultaat.verschillen)}")
    lines.append(f"  - Kritiek: {resultaat.aantal_kritieke_verschillen}")
    lines.append(f"  - Significant: {resultaat.aantal_significante_verschillen}")
    lines.append("")
    
    pensioen_totaal = resultaat.pensioen_resultaat.totaal_te_betalen_terug
    if pensioen_totaal >= Decimal("0"):
        lines.append(
            f"- **Pensioen-app totaal**: €{pensioen_totaal:,.2f} **bij te betalen**"
        )
    else:
        lines.append(
            f"- **Pensioen-app totaal**: €{abs(pensioen_totaal):,.2f} **terug te ontvangen**"
        )
    lines.append("")
    
    # Conclusies
    if resultaat.conclusies:
        lines.append("### Conclusies")
        lines.append("")
        for conclusie in resultaat.conclusies:
            lines.append(f"- {conclusie}")
        lines.append("")
    
    # Aanbevelingen
    if resultaat.aanbevelingen:
        lines.append("### 🎯 Aanbevelingen voor Pensioen-app")
        lines.append("")
        for idx, aanbeveling in enumerate(resultaat.aanbevelingen, 1):
            lines.append(f"{idx}. {aanbeveling}")
        lines.append("")
    
    lines.append("---")
    lines.append("")
    
    # Gedetailleerde verschillen per categorie
    lines.append("## 🔍 Gedetailleerde Verschillenanalyse")
    lines.append("")
    
    # Groepeer per categorie
    categorieën = {}
    for verschil in resultaat.verschillen:
        cat = verschil.categorie
        if cat not in categorieën:
            categorieën[cat] = []
        categorieën[cat].append(verschil)
    
    for categorie, verschillen in sorted(categorieën.items()):
        lines.append(f"### {categorie}")
        lines.append("")
        lines.append("| Onderdeel | dutch_tax | pensioen-app | Verschil | % | Ernst | Hypothese |")
        lines.append("|-----------|-----------|--------------|----------|---|-------|-----------|")
        
        for v in verschillen:
            dt_waarde = _fmt_bedrag(v.dutch_tax_waarde)
            pa_waarde = _fmt_bedrag(v.pensioen_app_waarde)
            verschil = _fmt_bedrag(v.verschil_absoluut, toon_plus=True)
            
            if v.verschil_procent is not None:
                pct = f"{v.verschil_procent:+.1f}%"
            else:
                pct = "—"
            
            ernst_emoji = {
                "KRITIEK": "🔴",
                "SIGNIFICANT": "🟡",
                "KLEIN": "🔵",
                "VERWAARLOOSBAAR": "⚪",
            }.get(v.ernst, "")
            
            # Escape pipe characters in hypothese
            hypothese = v.oorzaak_hypothese.replace("|", "\\|")
            
            lines.append(
                f"| {v.onderdeel} | {dt_waarde} | {pa_waarde} | {verschil} | "
                f"{pct} | {ernst_emoji} {v.ernst} | {hypothese} |"
            )
        
        lines.append("")
    
    # Pensioen-app berekening details
    lines.append("---")
    lines.append("")
    lines.append("## 💰 Pensioen-app Berekening Details")
    lines.append("")
    
    for persoon in resultaat.pensioen_resultaat.personen:
        lines.append(f"### {persoon.naam} ({persoon.persoon_id})")
        lines.append("")
        lines.append("**Box 1**")
        lines.append("")
        lines.append(f"- Bruto inkomen: €{persoon.bruto_inkomen:,.2f}")
        lines.append(f"- Waarvan arbeidsinkomen: €{persoon.arbeidsinkomen:,.2f}")
        lines.append(f"- Belasting voor kortingen: €{persoon.belasting_voor_kortingen:,.2f}")
        lines.append("")
        lines.append("**Heffingskortingen**")
        lines.append("")
        lines.append(f"- Algemene heffingskorting: €{persoon.ahk:,.2f}")
        lines.append(f"- Arbeidskorting: €{persoon.arbeidskorting:,.2f}")
        lines.append(f"- Ouderenkorting: €{persoon.ouderenkorting:,.2f}")
        lines.append(f"- **Totaal**: €{persoon.totale_heffingskorting:,.2f}")
        lines.append("")
        lines.append(f"- Netto belasting Box 1: €{persoon.netto_belasting_box1:,.2f}")
        lines.append(f"- Box 3 aandeel: €{persoon.box3_aandeel_heffing:,.2f}")
        lines.append(f"- **Totale belasting**: €{persoon.totale_belasting:,.2f}")
        lines.append("")
        lines.append(f"- Vooraf betaald: €{persoon.vooraf_betaald:,.2f}")
        
        if persoon.te_betalen_terug >= Decimal("0"):
            lines.append(f"- **Te betalen**: €{persoon.te_betalen_terug:,.2f}")
        else:
            lines.append(f"- **Terug te ontvangen**: €{abs(persoon.te_betalen_terug):,.2f}")
        lines.append("")
        
        if persoon.aannames:
            lines.append("**Aannames**")
            lines.append("")
            for aanname in persoon.aannames:
                lines.append(f"- {aanname}")
            lines.append("")
    
    # Box 3 huishoudniveau
    lines.append("### Box 3 (Huishouden)")
    lines.append("")
    lines.append(f"- Totaal vermogen: €{resultaat.pensioen_resultaat.box3_totaal_vermogen:,.2f}")
    lines.append(f"- Vrijstelling: €{resultaat.pensioen_resultaat.box3_vrijstelling:,.2f}")
    lines.append(
        f"- Belastbaar vermogen: €{resultaat.pensioen_resultaat.box3_belastbaar_vermogen:,.2f}"
    )
    lines.append(
        f"- Spaargeld fractie: {resultaat.pensioen_resultaat.box3_spaargeld_fractie:.1%}"
    )
    lines.append(f"- **Totale Box 3 heffing**: €{resultaat.pensioen_resultaat.box3_totale_heffing:,.2f}")
    lines.append("")
    
    # Metadata
    lines.append("---")
    lines.append("")
    lines.append("## ℹ️ Metadata")
    lines.append("")
    lines.append(f"- Belastingconfig gebruikt: {resultaat.pensioen_resultaat.config_gebruikt}")
    lines.append(f"- Heeft fiscaal partner: {resultaat.dutch_tax_data.heeft_fiscaal_partner}")
    lines.append(f"- Aantal kinderen: {resultaat.dutch_tax_data.aantal_kinderen}")
    lines.append("")
    
    # Disclaimer
    lines.append("---")
    lines.append("")
    lines.append("## ⚠️ Disclaimer")
    lines.append("")
    lines.append(
        "Dit is een **validatie tool** die dutch_tax (2025 submission) vergelijkt met "
        "pensioen-app berekening (2026 tarieven). Verschillen kunnen veroorzaakt worden door:"
    )
    lines.append("")
    lines.append("1. Tariefjaar verschil (2025 vs 2026)")
    lines.append("2. Ontbrekende features in pensioen-app (eigenwoningforfait, aftrekposten)")
    lines.append("3. Andere berekeningssystematiek (heffingskortingen, Box 3 forfaits)")
    lines.append("4. Data mapping issues")
    lines.append("")
    lines.append("**Deze tool is NIET bedoeld voor productiegebruik** — alleen voor interne validatie.")
    lines.append("")
    
    return "\n".join(lines)


def _fmt_bedrag(bedrag: Decimal, toon_plus: bool = False) -> str:
    """Formatteer bedrag als €X,XXX.XX."""
    prefix = ""
    if bedrag >= Decimal("0") and toon_plus:
        prefix = "+"
    elif bedrag < Decimal("0"):
        prefix = "-"
        bedrag = abs(bedrag)
    
    return f"{prefix}€{bedrag:,.2f}"


def schrijf_markdown_bestand(resultaat: VergelijkingsResultaat, uitvoer_pad: Path) -> None:
    """
    Schrijf markdown rapport naar bestand.
    
    Args:
        resultaat: VergelijkingsResultaat
        uitvoer_pad: Pad waar markdown bestand moet worden opgeslagen
    """
    markdown = genereer_markdown_rapport(resultaat)
    
    uitvoer_pad.parent.mkdir(parents=True, exist_ok=True)
    with open(uitvoer_pad, "w", encoding="utf-8") as f:
        f.write(markdown)


def genereer_excel_rapport(resultaat: VergelijkingsResultaat) -> bytes:
    """
    Genereer Excel rapport van vergelijking.
    
    Args:
        resultaat: VergelijkingsResultaat van vergelijking
    
    Returns:
        Excel bestand als bytes
    
    Raises:
        ImportError: Als openpyxl niet beschikbaar is
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        raise ImportError(
            "openpyxl is vereist voor Excel export. Installeer met: pip install openpyxl"
        ) from e
    
    wb = Workbook()
    
    # Sheet 1: Samenvatting
    ws_summary = wb.active
    ws_summary.title = "Samenvatting"
    
    ws_summary["A1"] = "Belastingvergelijking: dutch_tax vs pensioen-app"
    ws_summary["A1"].font = Font(bold=True, size=14)
    
    row = 3
    ws_summary[f"A{row}"] = "Huishouden ID"
    ws_summary[f"B{row}"] = resultaat.huishouden_id
    row += 1
    ws_summary[f"A{row}"] = "Submission jaar"
    ws_summary[f"B{row}"] = resultaat.submission_jaar
    row += 1
    ws_summary[f"A{row}"] = "Berekening jaar"
    ws_summary[f"B{row}"] = resultaat.jaar
    row += 1
    ws_summary[f"A{row}"] = "Gegenereerd"
    ws_summary[f"B{row}"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    row += 2
    
    ws_summary[f"A{row}"] = "Totaal verschillen"
    ws_summary[f"B{row}"] = len(resultaat.verschillen)
    row += 1
    ws_summary[f"A{row}"] = "Kritieke verschillen"
    ws_summary[f"B{row}"] = resultaat.aantal_kritieke_verschillen
    row += 1
    ws_summary[f"A{row}"] = "Significante verschillen"
    ws_summary[f"B{row}"] = resultaat.aantal_significante_verschillen
    row += 2
    
    ws_summary[f"A{row}"] = "Pensioen-app totaal"
    ws_summary[f"B{row}"] = float(resultaat.pensioen_resultaat.totaal_te_betalen_terug)
    ws_summary[f"B{row}"].number_format = "€#,##0.00"
    
    # Sheet 2: Verschillen
    ws_diff = wb.create_sheet("Verschillen")
    headers = [
        "Categorie",
        "Onderdeel",
        "dutch_tax",
        "pensioen-app",
        "Verschil",
        "Verschil %",
        "Ernst",
        "Hypothese",
    ]
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws_diff.cell(1, col_idx, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    for row_idx, verschil in enumerate(resultaat.verschillen, 2):
        ws_diff.cell(row_idx, 1, verschil.categorie)
        ws_diff.cell(row_idx, 2, verschil.onderdeel)
        
        dt_cell = ws_diff.cell(row_idx, 3, float(verschil.dutch_tax_waarde))
        dt_cell.number_format = "€#,##0.00"
        
        pa_cell = ws_diff.cell(row_idx, 4, float(verschil.pensioen_app_waarde))
        pa_cell.number_format = "€#,##0.00"
        
        diff_cell = ws_diff.cell(row_idx, 5, float(verschil.verschil_absoluut))
        diff_cell.number_format = "€#,##0.00"
        
        if verschil.verschil_procent is not None:
            pct_cell = ws_diff.cell(row_idx, 6, float(verschil.verschil_procent) / 100)
            pct_cell.number_format = "0.0%"
        else:
            ws_diff.cell(row_idx, 6, "—")
        
        ernst_cell = ws_diff.cell(row_idx, 7, verschil.ernst)
        if verschil.ernst == "KRITIEK":
            ernst_cell.fill = PatternFill(
                start_color="FF0000", end_color="FF0000", fill_type="solid"
            )
            ernst_cell.font = Font(color="FFFFFF", bold=True)
        elif verschil.ernst == "SIGNIFICANT":
            ernst_cell.fill = PatternFill(
                start_color="FFFF00", end_color="FFFF00", fill_type="solid"
            )
        
        ws_diff.cell(row_idx, 8, verschil.oorzaak_hypothese)
    
    # Auto-size kolommen
    for col in range(1, 9):
        ws_diff.column_dimensions[get_column_letter(col)].width = 20
    ws_diff.column_dimensions["H"].width = 50  # Hypothese kolom breder
    
    # Sheet 3: Aanbevelingen
    ws_rec = wb.create_sheet("Aanbevelingen")
    ws_rec["A1"] = "Aanbevelingen voor Pensioen-app"
    ws_rec["A1"].font = Font(bold=True, size=12)
    
    for row_idx, aanbeveling in enumerate(resultaat.aanbevelingen, 3):
        ws_rec.cell(row_idx, 1, f"{row_idx - 2}.")
        ws_rec.cell(row_idx, 2, aanbeveling)
    
    ws_rec.column_dimensions["B"].width = 80
    
    # Schrijf naar BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def schrijf_excel_bestand(resultaat: VergelijkingsResultaat, uitvoer_pad: Path) -> None:
    """
    Schrijf Excel rapport naar bestand.
    
    Args:
        resultaat: VergelijkingsResultaat
        uitvoer_pad: Pad waar Excel bestand moet worden opgeslagen
    """
    excel_bytes = genereer_excel_rapport(resultaat)
    
    uitvoer_pad.parent.mkdir(parents=True, exist_ok=True)
    with open(uitvoer_pad, "wb") as f:
        f.write(excel_bytes)
